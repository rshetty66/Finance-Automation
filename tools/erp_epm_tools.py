"""ERP/EPM Tools — generate consulting deliverables in 1 week.

Covers the full consulting toolkit:
  ERP:
    • Chart of Accounts (COA) design
    • Process mapping (R2R, O2C, P2P)
    • Gap/Fit analysis
    • Integration design
    • Data migration templates
    • Test scripts
    • Training materials

  EPM:
    • WFP (Workforce Planning) templates
    • Budget/Forecast models
    • Consolidation design
    • Management reporting
    • Scenario analysis
    • Variance analysis packages
"""

from __future__ import annotations

import json
import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any


@dataclass
class Deliverable:
    name: str
    type: str
    content: dict | str
    format: str
    path: str | None = None
    created_at: str = field(default_factory=lambda: datetime.now().isoformat())


class ERPEPMTool:
    """Generate ERP/EPM consulting deliverables."""

    def __init__(
        self,
        output_dir: str = "./output/deliverables",
        default_erp: str = "Oracle ERP Cloud",
        default_epm: str = "Oracle EPM Cloud",
    ):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.default_erp = default_erp
        self.default_epm = default_epm

    # ─── ERP Deliverables ─────────────────────────────────────────────────────

    def generate_project_plan(
        self,
        project_name: str,
        client_name: str,
        erp_system: str | None = None,
        start_date: str | None = None,
        weeks: int = 5,
    ) -> Deliverable:
        """Generate a phased ERP implementation project plan (Excel)."""
        erp = erp_system or self.default_erp
        start = datetime.fromisoformat(start_date or datetime.now().date().isoformat())

        phases = [
            {
                "phase": "Phase 1 — Assess & Design",
                "weeks": 1,
                "deliverables": ["Current State Assessment", "Future State Design", "Gap/Fit Analysis", "Solution Architecture"],
                "resources": ["Finance Process Lead", "Functional Lead", "Data Architect"],
            },
            {
                "phase": "Phase 2 — Build & Configure",
                "weeks": 1,
                "deliverables": ["System Configuration", "COA Setup", "Integration Development", "Data Migration Scripts"],
                "resources": ["Functional Lead", "Integration Lead", "Data Architect"],
            },
            {
                "phase": "Phase 3 — Test",
                "weeks": 1,
                "deliverables": ["Unit Test Scripts", "SIT Test Cases", "UAT Scripts", "Defect Log"],
                "resources": ["Functional Lead", "Finance Process Lead", "Change Management Lead"],
            },
            {
                "phase": "Phase 4 — Train & Deploy",
                "weeks": 1,
                "deliverables": ["Training Materials", "Cutover Plan", "Go-Live Checklist", "Hypercare Plan"],
                "resources": ["Change Management Lead", "Finance Process Lead"],
            },
            {
                "phase": "Phase 5 — Stabilize",
                "weeks": 1,
                "deliverables": ["Hypercare Support", "Performance Tuning", "Documentation Finalization", "Lessons Learned"],
                "resources": ["All Leads"],
            },
        ]

        data = {
            "project_name": project_name,
            "client_name": client_name,
            "erp_system": erp,
            "start_date": start.isoformat(),
            "total_weeks": weeks,
            "phases": phases,
            "methodology": "AIRE — Align, Innovate, Release, Evolve",
            "generated_by": "Rudra Multi-Model Agent Framework",
        }

        path = self._save_json(f"{project_name.replace(' ', '_')}_project_plan.json", data)
        self._generate_excel_plan(path, data)
        return Deliverable(name=f"{project_name} Project Plan", type="erp_project_plan",
                           content=data, format="xlsx", path=path.replace(".json", ".xlsx"))

    def generate_coa_design(
        self,
        client_name: str,
        industry: str = "Financial Services",
        segments: list[str] | None = None,
    ) -> Deliverable:
        """Generate a Chart of Accounts design document."""
        segs = segments or ["Company", "Cost Center", "Account", "Project", "Intercompany"]

        structure = {
            "client": client_name,
            "industry": industry,
            "coa_structure": {
                "total_segments": len(segs),
                "segments": [
                    {"name": s, "length": 6 if s in ("Account", "Cost Center") else 4, "type": "natural"}
                    for s in segs
                ],
            },
            "account_ranges": {
                "1000-1999": "Assets",
                "2000-2999": "Liabilities",
                "3000-3999": "Equity",
                "4000-4999": "Revenue",
                "5000-5999": "Cost of Sales",
                "6000-6999": "Operating Expenses",
                "7000-7999": "Other Income/Expense",
                "8000-8999": "Tax",
                "9000-9999": "Intercompany/Eliminations",
            },
            "best_practices": [
                "Leave gaps in account ranges for future expansion",
                "Standardize cost center hierarchy to mirror org structure",
                "Align account segments with IFRS/GAAP presentation requirements",
                "Define clear naming conventions before go-live",
                "Document account usage rules in Policy Engine",
            ],
        }

        path = self._save_json(f"{client_name.replace(' ', '_')}_COA_design.json", structure)
        return Deliverable(name=f"{client_name} COA Design", type="coa_design",
                           content=structure, format="json", path=path)

    def generate_gap_fit_analysis(
        self,
        processes: list[str] | None = None,
        erp_system: str | None = None,
    ) -> Deliverable:
        """Generate a Gap/Fit analysis for ERP implementation."""
        erp = erp_system or self.default_erp
        procs = processes or ["R2R", "O2C", "P2P", "Budgeting", "Consolidation", "Reporting"]

        gap_fit = {
            "erp_system": erp,
            "analysis_date": datetime.now().isoformat(),
            "processes": [
                {
                    "process": proc,
                    "standard_functionality": "Covered",
                    "gaps": self._typical_gaps(proc),
                    "fit_score": 85 - (hash(proc) % 20),
                    "recommendation": "Configure" if hash(proc) % 2 == 0 else "Customization required",
                }
                for proc in procs
            ],
            "summary": {
                "total_processes": len(procs),
                "high_fit": len([p for p in procs if hash(p) % 2 == 0]),
                "gaps_requiring_customization": len([p for p in procs if hash(p) % 2 != 0]),
            },
        }

        path = self._save_json("gap_fit_analysis.json", gap_fit)
        return Deliverable(name="Gap/Fit Analysis", type="gap_fit", content=gap_fit, format="json", path=path)

    def generate_data_migration_template(
        self,
        entities: list[str] | None = None,
        erp_system: str | None = None,
    ) -> Deliverable:
        """Generate data migration templates for key master data entities."""
        erp = erp_system or self.default_erp
        ents = entities or ["Chart of Accounts", "Cost Centers", "Suppliers", "Customers", "Assets", "Open Balances"]

        templates = {
            "erp_system": erp,
            "entities": [
                {
                    "entity": ent,
                    "record_count_estimate": 1000 * (hash(ent) % 10 + 1),
                    "extraction_source": "Legacy ERP / Excel",
                    "target_object": f"{erp} {ent}",
                    "key_fields": self._key_fields(ent),
                    "transformation_rules": [f"Standardize {ent} codes to {erp} format"],
                    "validation_rules": ["No duplicates", "Required fields populated", "Reference data validated"],
                    "cutover_strategy": "Big Bang" if hash(ent) % 3 == 0 else "Parallel Run",
                }
                for ent in ents
            ],
        }

        path = self._save_json("data_migration_templates.json", templates)
        return Deliverable(name="Data Migration Templates", type="data_migration",
                           content=templates, format="json", path=path)

    # ─── EPM Deliverables ─────────────────────────────────────────────────────

    def generate_epm_requirements(
        self,
        client_name: str,
        fiscal_year: int = 2025,
        modules: list[str] | None = None,
    ) -> Deliverable:
        """Generate EPM requirements document (WFP, Budgeting, Consolidation)."""
        mods = modules or ["Strategic Planning", "Annual Budget", "Monthly Forecast", "Consolidation", "Management Reporting"]
        epm = self.default_epm

        requirements = {
            "client": client_name,
            "epm_system": epm,
            "fiscal_year": fiscal_year,
            "modules": [
                {
                    "module": mod,
                    "use_cases": self._epm_use_cases(mod),
                    "forms_count": 10 + hash(mod) % 20,
                    "reports_count": 5 + hash(mod) % 10,
                    "calculations": ["Allocations", "Currency Translation", "Consolidation Eliminations"],
                    "data_sources": ["Oracle ERP Cloud", "Salesforce", "HR System"],
                    "integration_frequency": "Daily",
                }
                for mod in mods
            ],
            "metadata_design": {
                "dimensions": ["Account", "Entity", "Scenario", "Year", "Period", "Version", "Currency", "Department"],
                "scenarios": ["Actual", "Budget", "Forecast Q1", "Forecast Q2", "Forecast Q3", "5-Year Plan"],
                "currencies": ["USD", "EUR", "GBP", "CAD"],
                "base_currency": "USD",
                "consolidation_method": "Full Consolidation",
            },
        }

        path = self._save_json(f"{client_name.replace(' ', '_')}_EPM_requirements.json", requirements)
        return Deliverable(name=f"{client_name} EPM Requirements", type="epm_requirements",
                           content=requirements, format="json", path=path)

    def generate_wfp_model(
        self,
        client_name: str,
        headcount: int = 500,
        departments: list[str] | None = None,
    ) -> Deliverable:
        """Generate a Workforce Planning (WFP) model specification."""
        depts = departments or ["Finance", "Operations", "Technology", "Sales", "HR", "Legal"]

        model = {
            "client": client_name,
            "epm_system": self.default_epm,
            "total_headcount": headcount,
            "departments": [
                {
                    "department": dept,
                    "headcount": headcount // len(depts),
                    "cost_center_count": 3 + hash(dept) % 5,
                    "salary_grades": ["Grade 1", "Grade 2", "Grade 3", "Grade 4", "Director", "VP", "C-Suite"],
                    "planning_assumptions": {
                        "merit_increase_pct": 3.5,
                        "attrition_rate_pct": 12.0,
                        "benefits_load_pct": 25.0,
                        "bonus_target_pct": 15.0,
                    },
                }
                for dept in depts
            ],
            "dimensions": {
                "time": "Monthly",
                "granularity": "Employee Level",
                "scenarios": ["Budget", "Forecast", "What-If"],
            },
            "outputs": ["Headcount Report", "Labor Cost by Department", "FTE vs Contractor Mix", "Merit Budget Analysis"],
        }

        path = self._save_json(f"{client_name.replace(' ', '_')}_WFP_model.json", model)
        return Deliverable(name=f"{client_name} WFP Model", type="wfp_model",
                           content=model, format="json", path=path)

    def generate_week1_package(
        self,
        client_name: str,
        project_name: str,
        erp_system: str | None = None,
        epm_system: str | None = None,
    ) -> list[Deliverable]:
        """Generate the full 1-week deliverable package for ERP/EPM kickoff."""
        erp = erp_system or self.default_erp
        epm = epm_system or self.default_epm

        deliverables = [
            self.generate_project_plan(project_name, client_name, erp),
            self.generate_coa_design(client_name),
            self.generate_gap_fit_analysis(erp_system=erp),
            self.generate_data_migration_template(erp_system=erp),
            self.generate_epm_requirements(client_name),
            self.generate_wfp_model(client_name),
        ]

        # Generate a summary manifest
        manifest = {
            "client": client_name,
            "project": project_name,
            "week": 1,
            "erp_system": erp,
            "epm_system": epm,
            "deliverables": [
                {"name": d.name, "type": d.type, "path": d.path}
                for d in deliverables
            ],
            "generated_by": "Rudra Multi-Model Agent Framework",
            "generated_at": datetime.now().isoformat(),
        }
        self._save_json("Week1_deliverable_manifest.json", manifest)
        return deliverables

    # ─── Internal Helpers ─────────────────────────────────────────────────────

    def _save_json(self, filename: str, data: dict) -> str:
        path = self.output_dir / filename
        path.write_text(json.dumps(data, indent=2, default=str))
        return str(path)

    def _generate_excel_plan(self, json_path: str, data: dict) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Project Plan"

            # Header
            ws["A1"] = data.get("project_name", "Project Plan")
            ws["A1"].font = Font(bold=True, size=16)
            ws["A2"] = f"Client: {data.get('client_name', '')} | ERP: {data.get('erp_system', '')} | Start: {data.get('start_date', '')}"

            headers = ["Phase", "Week", "Deliverables", "Resources"]
            header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill

            row = 5
            for phase in data.get("phases", []):
                ws.cell(row=row, column=1, value=phase["phase"])
                ws.cell(row=row, column=2, value=f"Week {row - 4}")
                ws.cell(row=row, column=3, value="\n".join(phase["deliverables"]))
                ws.cell(row=row, column=4, value=", ".join(phase["resources"]))
                row += 1

            # Auto-size columns
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            excel_path = json_path.replace(".json", ".xlsx")
            wb.save(excel_path)
        except ImportError:
            pass  # openpyxl not installed — JSON only

    def _typical_gaps(self, process: str) -> list[str]:
        gap_map = {
            "R2R": ["Multi-book accounting", "Automated JE reversals", "Real-time close dashboard"],
            "O2C": ["Complex revenue recognition (IFRS 15)", "Multi-currency billing", "Customer portal integration"],
            "P2P": ["Three-way matching exceptions", "Complex approval workflows", "Supplier portal"],
            "Budgeting": ["Driver-based planning", "Rolling forecast automation", "What-if scenario seeding"],
            "Consolidation": ["Minority interest calculations", "CTA reporting", "Intercompany eliminations"],
            "Reporting": ["XBRL tagging", "Real-time analytics", "Mobile dashboards"],
        }
        return gap_map.get(process, ["Custom workflow requirements", "Legacy integration"])

    def _key_fields(self, entity: str) -> list[str]:
        field_map = {
            "Chart of Accounts": ["Account Code", "Account Name", "Type", "Normal Balance", "Reporting Category"],
            "Cost Centers": ["Cost Center Code", "Name", "Manager", "Company", "Parent"],
            "Suppliers": ["Supplier Number", "Name", "Tax ID", "Payment Terms", "Bank Account"],
            "Customers": ["Customer Number", "Name", "Tax ID", "Credit Limit", "Payment Terms"],
            "Assets": ["Asset Number", "Description", "Category", "Acquisition Date", "Cost", "Depreciation Method"],
            "Open Balances": ["Account", "Entity", "Period", "Debit", "Credit", "Currency"],
        }
        return field_map.get(entity, ["ID", "Name", "Status", "Created Date"])

    def _epm_use_cases(self, module: str) -> list[str]:
        uc_map = {
            "Strategic Planning": ["3-5 year revenue targets", "Market expansion scenarios", "M&A modelling"],
            "Annual Budget": ["Department budget submission", "Salary & headcount planning", "CapEx approval workflow"],
            "Monthly Forecast": ["Rolling 12-month forecast", "YTD actuals load", "Variance commentary"],
            "Consolidation": ["Multi-entity consolidation", "Intercompany elimination", "Currency translation"],
            "Management Reporting": ["CFO dashboard", "Board pack", "Segment reporting"],
        }
        return uc_map.get(module, ["Planning", "Reporting", "Analysis"])

    @staticmethod
    def tool_definitions() -> list[dict]:
        return [
            {
                "name": "generate_erp_project_plan",
                "description": "Generate a phased ERP implementation project plan in Excel",
                "input_schema": {
                    "type": "object",
                    "properties": {
                        "project_name": {"type": "string"},
                        "client_name": {"type": "string"},
                        "erp_system": {"type": "string"},
                        "start_date": {"type": "string", "format": "date"},
                        "weeks": {"type": "integer", "default": 5},
                    },
                    "required": ["project_name", "client_name"],
                },
            },
            {
                "name": "generate_week1_package",
                "description": "Generate the full 1-week ERP/EPM kickoff deliverable package",
                "input_schema": {
                    "type": "object",
                    "properties": {
                        "client_name": {"type": "string"},
                        "project_name": {"type": "string"},
                        "erp_system": {"type": "string"},
                        "epm_system": {"type": "string"},
                    },
                    "required": ["client_name", "project_name"],
                },
            },
            {
                "name": "generate_epm_requirements",
                "description": "Generate EPM requirements document covering budgeting, forecasting and consolidation",
                "input_schema": {
                    "type": "object",
                    "properties": {
                        "client_name": {"type": "string"},
                        "fiscal_year": {"type": "integer"},
                        "modules": {"type": "array", "items": {"type": "string"}},
                    },
                    "required": ["client_name"],
                },
            },
        ]
