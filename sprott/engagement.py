"""
Sprott Finance Transformation Engagement.

Pipeline:
  input files → extract content → Claude analysis → HTML deliverable
"""

from __future__ import annotations

import json
import os
import textwrap
from datetime import date
from pathlib import Path
from typing import Any

from anthropic import Anthropic
from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------------------------------------
# File content extraction
# ---------------------------------------------------------------------------

def _extract_excel(path: Path) -> str:
    """Return a text summary of an Excel workbook (all sheets, first 40 rows each)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = [f"Excel file: {path.name}"]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"\n=== Sheet: {sheet_name} ===")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 40:
                    lines.append(f"  ... ({ws.max_row - 40} more rows)")
                    break
                non_null = [str(c) for c in row if c is not None]
                if non_null:
                    lines.append("  " + " | ".join(non_null))
        return "\n".join(lines)
    except Exception as exc:
        return f"[Excel extraction error: {exc}]"


def _extract_pdf(path: Path) -> str:
    """Return text content of a PDF (first 8 pages)."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        pages = reader.pages[:8]
        parts = [f"PDF file: {path.name}"]
        for i, page in enumerate(pages):
            text = page.extract_text() or ""
            parts.append(f"\n--- Page {i + 1} ---\n{text[:2000]}")
        return "\n".join(parts)
    except Exception as exc:
        return f"[PDF extraction error: {exc}]"


def _extract_csv(path: Path) -> str:
    """Return first 50 rows of a CSV."""
    try:
        lines = path.read_text(errors="replace").splitlines()
        header = "\n".join(lines[:50])
        note = f"\n... ({len(lines) - 50} more rows)" if len(lines) > 50 else ""
        return f"CSV file: {path.name}\n{header}{note}"
    except Exception as exc:
        return f"[CSV extraction error: {exc}]"


def extract_file_content(file_path: str) -> str:
    """Dispatch to the right extractor based on file extension."""
    path = Path(file_path)
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        return _extract_excel(path)
    if ext == ".pdf":
        return _extract_pdf(path)
    if ext == ".csv":
        return _extract_csv(path)
    # Fallback: plain text
    try:
        return path.read_text(errors="replace")[:8000]
    except Exception as exc:
        return f"[Unreadable file: {exc}]"


# ---------------------------------------------------------------------------
# Claude calls
# ---------------------------------------------------------------------------

SPROTT_CONTEXT = """
Sprott Inc. is a Toronto-based asset management company specializing in precious metals,
energy transition, and natural resource investments. They manage mutual funds, ETFs,
closed-end funds, and managed accounts with AUM in the range of $30–35 billion USD.

Key characteristics:
- Multi-entity structure: holding company + multiple fund entities + management company
- Regulated environment: SEC, OSC, NI 81-102 (Canadian mutual funds), Sarbanes-Oxley
- Finance processes: fund accounting, NAV calculation, management fee billing,
  investor reporting, regulatory reporting (Form N-CEN, N-PORT, SEDAR),
  intercompany allocations, treasury/cash management
- Technology landscape: likely mix of older fund admin systems, Excel-heavy close,
  manual reconciliations, minimal ERP integration
- Growth context: AUM growth via acquisitions (Sprott has acquired multiple fund
  families), driving need for finance transformation to scale operations
"""

ANALYSIS_SYSTEM_PROMPT = f"""You are a senior finance transformation consultant with deep expertise
in asset management, fund accounting, and ERP/EPM implementations for investment firms.

Client context:
{SPROTT_CONTEXT}

You are analyzing input files from Sprott to produce a Finance Transformation assessment.
Extract key findings, pain points, process inefficiencies, and data quality issues.
Output as structured JSON only — no prose outside the JSON.
"""

ROADMAP_SYSTEM_PROMPT = f"""You are a principal-level finance transformation consultant specializing in
asset management firms. You produce board-ready consulting deliverables.

Client context:
{SPROTT_CONTEXT}

Given a set of diagnostic findings, produce a comprehensive Finance Transformation
Roadmap and Project Plan for Sprott.

IMPORTANT: Be concise. Use short strings (under 15 words each). Strict limits:
- transformation_themes: exactly 3 items
- current_state_findings: exactly 5 items
- benchmark_gaps: exactly 4 items
- roadmap_phases: exactly 4 phases, each with exactly 2 workstreams, each workstream with exactly 3 activities, key_deliverables has 3 items, dependencies has 2 items
- quick_wins: exactly 4 items
- risks: exactly 3 items
- next_steps: exactly 4 items
- investment_summary.key_value_drivers: exactly 4 items

Output as structured JSON matching this exact schema:
{{
  "executive_summary": "2-3 sentence summary",
  "transformation_themes": [
    {{"theme": str, "description": str, "priority": "High|Medium|Low"}}
  ],
  "current_state_findings": [
    {{"area": str, "finding": str, "impact": str, "severity": "Critical|High|Medium"}}
  ],
  "benchmark_gaps": [
    {{"metric": str, "sprott_estimate": str, "top_quartile": str, "gap": str}}
  ],
  "roadmap_phases": [
    {{
      "phase": int,
      "name": str,
      "duration": str,
      "quarter": str,
      "workstreams": [
        {{"name": str, "activities": [str, str, str], "owner": str, "milestone": str}}
      ],
      "key_deliverables": [str, str, str],
      "dependencies": [str, str]
    }}
  ],
  "investment_summary": {{
    "total_investment": str,
    "annual_savings": str,
    "fte_reduction": str,
    "payback_period": str,
    "npv_3yr": str,
    "key_value_drivers": [str, str, str, str]
  }},
  "quick_wins": [
    {{"initiative": str, "timeline": str, "savings": str, "effort": "Low|Medium|High"}}
  ],
  "risks": [
    {{"risk": str, "likelihood": "High|Medium|Low", "mitigation": str}}
  ],
  "next_steps": [
    {{"action": str, "owner": str, "due": str}}
  ]
}}
"""


def _call_claude(system: str, user: str, model: str = "claude-opus-4-6", max_tokens: int = 4096) -> str:
    client = Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    response = client.messages.create(
        model=model,
        max_tokens=max_tokens,
        system=system,
        messages=[{"role": "user", "content": user}],
    )
    return response.content[0].text


def _parse_json(raw: str) -> dict[str, Any]:
    """Parse JSON from Claude response, stripping markdown fences."""
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()
    # Try direct parse first
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        # Find the outermost { ... } and try to parse that
        start = raw.find("{")
        end = raw.rfind("}")
        if start != -1 and end != -1:
            return json.loads(raw[start:end + 1])
        raise


def analyze_inputs(file_contents: list[str]) -> dict[str, Any]:
    """Step 1 — extract structured findings from raw file content."""
    combined = "\n\n".join(file_contents)
    prompt = f"""Analyze these Sprott finance files and extract diagnostic findings.

FILE CONTENT:
{combined[:12000]}

Return JSON with this shape:
{{
  "data_quality_issues": [{{"area": str, "issue": str, "volume": str}}],
  "process_inefficiencies": [{{"process": str, "pain_point": str, "estimated_effort": str}}],
  "technology_gaps": [{{"system": str, "gap": str, "risk": str}}],
  "close_cycle_observations": [{{"observation": str, "benchmark_gap": str}}],
  "key_metrics_found": [{{"metric": str, "value": str, "context": str}}],
  "summary": str
}}
"""
    raw = _call_claude(ANALYSIS_SYSTEM_PROMPT, prompt, model="claude-sonnet-4-6", max_tokens=8096)
    return _parse_json(raw)


def generate_roadmap(findings: dict[str, Any]) -> dict[str, Any]:
    """Step 2 — generate roadmap JSON from findings."""
    prompt = f"""Based on these diagnostic findings for Sprott, generate the Finance Transformation Roadmap.

FINDINGS:
{json.dumps(findings, indent=2)}

Today is {date.today().isoformat()}. Phase quarters should start from Q2 2026.
Return the full roadmap JSON as specified in your instructions.
"""
    raw = _call_claude(ROADMAP_SYSTEM_PROMPT, prompt, model="claude-sonnet-4-6", max_tokens=4096)
    return _parse_json(raw)


# ---------------------------------------------------------------------------
# HTML rendering
# ---------------------------------------------------------------------------

_SEVERITY_COLOR = {"Critical": "#dc2626", "High": "#ea580c", "Medium": "#d97706"}
_PRIORITY_COLOR = {"High": "#dc2626", "Medium": "#d97706", "Low": "#16a34a"}
_PHASE_COLORS = ["#1e3a5f", "#1d4ed8", "#0891b2", "#059669", "#7c3aed"]


def _badge(text: str, color: str) -> str:
    return (
        f'<span style="background:{color};color:#fff;padding:2px 8px;'
        f'border-radius:4px;font-size:11px;font-weight:600">{text}</span>'
    )


def render_html(findings: dict[str, Any], roadmap: dict[str, Any], input_files: list[str]) -> str:
    today = date.today().strftime("%B %d, %Y")
    file_names = ", ".join(Path(f).name for f in input_files)

    # --- Executive Summary ---
    exec_summary = roadmap.get("executive_summary", "")

    # --- Transformation Themes ---
    themes_html = ""
    for t in roadmap.get("transformation_themes", []):
        color = _PRIORITY_COLOR.get(t.get("priority", "Medium"), "#6b7280")
        themes_html += f"""
        <div style="background:#f8fafc;border-left:4px solid {color};padding:12px 16px;margin-bottom:10px;border-radius:0 6px 6px 0">
          <div style="font-weight:700;color:#1e293b">{t['theme']} {_badge(t.get('priority',''), color)}</div>
          <div style="color:#475569;margin-top:4px;font-size:14px">{t['description']}</div>
        </div>"""

    # --- Current State Findings ---
    findings_html = ""
    for f in roadmap.get("current_state_findings", []):
        color = _SEVERITY_COLOR.get(f.get("severity", "Medium"), "#6b7280")
        findings_html += f"""
        <tr>
          <td style="font-weight:600;color:#1e293b;padding:10px 12px;border-bottom:1px solid #e2e8f0">{f['area']}</td>
          <td style="color:#475569;padding:10px 12px;border-bottom:1px solid #e2e8f0">{f['finding']}</td>
          <td style="color:#475569;padding:10px 12px;border-bottom:1px solid #e2e8f0">{f['impact']}</td>
          <td style="padding:10px 12px;border-bottom:1px solid #e2e8f0">{_badge(f.get('severity',''), color)}</td>
        </tr>"""

    # --- Benchmark Gaps ---
    benchmarks_html = ""
    for b in roadmap.get("benchmark_gaps", []):
        benchmarks_html += f"""
        <tr>
          <td style="font-weight:600;color:#1e293b;padding:10px 12px;border-bottom:1px solid #e2e8f0">{b['metric']}</td>
          <td style="text-align:center;color:#dc2626;font-weight:700;padding:10px 12px;border-bottom:1px solid #e2e8f0">{b.get('sprott_estimate','—')}</td>
          <td style="text-align:center;color:#16a34a;font-weight:700;padding:10px 12px;border-bottom:1px solid #e2e8f0">{b.get('top_quartile','—')}</td>
          <td style="text-align:center;color:#d97706;font-weight:600;padding:10px 12px;border-bottom:1px solid #e2e8f0">{b.get('gap','—')}</td>
        </tr>"""

    # --- Roadmap Phases (visual timeline) ---
    phases = roadmap.get("roadmap_phases", [])
    timeline_html = ""
    for ph in phases:
        idx = (ph.get("phase", 1) - 1) % len(_PHASE_COLORS)
        color = _PHASE_COLORS[idx]
        workstreams_html = ""
        for ws in ph.get("workstreams", []):
            acts = "".join(f"<li>{a}</li>" for a in ws.get("activities", []))
            workstreams_html += f"""
            <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;padding:12px;margin-bottom:8px">
              <div style="font-weight:700;color:{color};font-size:13px">{ws['name']}
                <span style="font-weight:400;color:#64748b;float:right">{ws.get('owner','')}</span>
              </div>
              <ul style="margin:6px 0 0 16px;color:#475569;font-size:13px">{acts}</ul>
              <div style="margin-top:6px;font-size:12px;color:#7c3aed">
                ✓ Milestone: {ws.get('milestone','')}
              </div>
            </div>"""

        deliverables = "".join(f"<li>{d}</li>" for d in ph.get("key_deliverables", []))
        deps = "".join(f"<span style='background:#fef3c7;color:#92400e;padding:2px 6px;border-radius:3px;font-size:11px;margin-right:4px'>{d}</span>"
                       for d in ph.get("dependencies", []))

        timeline_html += f"""
        <div style="border:2px solid {color};border-radius:10px;margin-bottom:24px;overflow:hidden">
          <div style="background:{color};color:#fff;padding:14px 20px;display:flex;justify-content:space-between;align-items:center">
            <div>
              <span style="font-size:13px;opacity:0.8">Phase {ph.get('phase','')}</span>
              <div style="font-size:20px;font-weight:800;margin-top:2px">{ph.get('name','')}</div>
            </div>
            <div style="text-align:right">
              <div style="font-size:13px;opacity:0.8">{ph.get('quarter','')}</div>
              <div style="font-size:16px;font-weight:700">{ph.get('duration','')}</div>
            </div>
          </div>
          <div style="padding:16px 20px">
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">
              <div>
                <div style="font-weight:700;color:#1e293b;margin-bottom:8px;font-size:14px">WORKSTREAMS</div>
                {workstreams_html}
              </div>
              <div>
                <div style="font-weight:700;color:#1e293b;margin-bottom:8px;font-size:14px">KEY DELIVERABLES</div>
                <ul style="color:#475569;font-size:13px;margin:0 0 0 16px">{deliverables}</ul>
                {f'<div style="margin-top:12px"><div style="font-weight:700;color:#1e293b;font-size:13px;margin-bottom:6px">DEPENDENCIES</div>{deps}</div>' if deps else ''}
              </div>
            </div>
          </div>
        </div>"""

    # --- Investment Summary ---
    inv = roadmap.get("investment_summary", {})
    inv_cards = ""
    inv_metrics = [
        ("Total Investment", inv.get("total_investment", "—"), "#1e3a5f"),
        ("Annual Savings", inv.get("annual_savings", "—"), "#16a34a"),
        ("FTE Reduction", inv.get("fte_reduction", "—"), "#0891b2"),
        ("Payback Period", inv.get("payback_period", "—"), "#7c3aed"),
        ("3-Year NPV", inv.get("npv_3yr", "—"), "#059669"),
    ]
    for label, value, color in inv_metrics:
        inv_cards += f"""
        <div style="background:#f8fafc;border-top:4px solid {color};border-radius:8px;padding:16px;text-align:center">
          <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:0.05em">{label}</div>
          <div style="font-size:22px;font-weight:800;color:{color};margin-top:6px">{value}</div>
        </div>"""

    value_drivers = "".join(f"<li>{d}</li>" for d in inv.get("key_value_drivers", []))

    # --- Quick Wins ---
    qw_html = ""
    for qw in roadmap.get("quick_wins", []):
        effort_color = {"Low": "#16a34a", "Medium": "#d97706", "High": "#dc2626"}.get(qw.get("effort", "Medium"), "#6b7280")
        qw_html += f"""
        <tr>
          <td style="font-weight:600;color:#1e293b;padding:10px 12px;border-bottom:1px solid #e2e8f0">{qw['initiative']}</td>
          <td style="color:#475569;padding:10px 12px;border-bottom:1px solid #e2e8f0">{qw.get('timeline','')}</td>
          <td style="color:#16a34a;font-weight:600;padding:10px 12px;border-bottom:1px solid #e2e8f0">{qw.get('savings','')}</td>
          <td style="padding:10px 12px;border-bottom:1px solid #e2e8f0">{_badge(qw.get('effort',''), effort_color)}</td>
        </tr>"""

    # --- Risks ---
    risks_html = ""
    for r in roadmap.get("risks", []):
        color = _SEVERITY_COLOR.get(r.get("likelihood", "Medium"), "#6b7280")
        risks_html += f"""
        <tr>
          <td style="font-weight:600;color:#1e293b;padding:10px 12px;border-bottom:1px solid #e2e8f0">{r['risk']}</td>
          <td style="padding:10px 12px;border-bottom:1px solid #e2e8f0">{_badge(r.get('likelihood',''), color)}</td>
          <td style="color:#475569;padding:10px 12px;border-bottom:1px solid #e2e8f0">{r.get('mitigation','')}</td>
        </tr>"""

    # --- Next Steps ---
    next_html = ""
    for i, step in enumerate(roadmap.get("next_steps", []), 1):
        next_html += f"""
        <div style="display:flex;align-items:flex-start;gap:12px;padding:12px;border-bottom:1px solid #e2e8f0">
          <div style="background:#1e3a5f;color:#fff;border-radius:50%;width:28px;height:28px;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:13px;flex-shrink:0">{i}</div>
          <div style="flex:1">
            <div style="font-weight:600;color:#1e293b">{step['action']}</div>
            <div style="font-size:13px;color:#64748b;margin-top:2px">Owner: {step.get('owner','')} · Due: {step.get('due','')}</div>
          </div>
        </div>"""

    # --- Assemble full page ---
    section_style = "background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.08);padding:28px;margin-bottom:28px"
    h2_style = "font-size:18px;font-weight:800;color:#1e3a5f;margin:0 0 18px 0;padding-bottom:10px;border-bottom:2px solid #e2e8f0;text-transform:uppercase;letter-spacing:0.04em"
    table_style = "width:100%;border-collapse:collapse;font-size:14px"
    th_style = "background:#f1f5f9;color:#475569;font-size:11px;text-transform:uppercase;letter-spacing:0.05em;padding:10px 12px;text-align:left"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Sprott — Finance Transformation Roadmap</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
         background:#f1f5f9; color:#1e293b; margin:0; padding:0; }}
  @media print {{ body {{ background:#fff; }} }}
</style>
</head>
<body>

<!-- HEADER -->
<div style="background:linear-gradient(135deg,#1e3a5f 0%,#1d4ed8 100%);color:#fff;padding:40px 48px">
  <div style="max-width:1100px;margin:0 auto">
    <div style="display:flex;justify-content:space-between;align-items:flex-start">
      <div>
        <div style="font-size:13px;opacity:0.7;text-transform:uppercase;letter-spacing:0.1em">Confidential · Finance Transformation</div>
        <h1 style="margin:8px 0 4px;font-size:36px;font-weight:900">Sprott Inc.</h1>
        <div style="font-size:20px;opacity:0.9;font-weight:300">Finance Transformation Roadmap &amp; Project Plan</div>
      </div>
      <div style="text-align:right;opacity:0.8;font-size:13px">
        <div style="font-weight:700;font-size:16px">{today}</div>
        <div style="margin-top:4px">Source files: {file_names}</div>
        <div style="margin-top:4px">Prepared by: Rudra AI · Confidential</div>
      </div>
    </div>
  </div>
</div>

<div style="max-width:1100px;margin:32px auto;padding:0 24px">

  <!-- EXECUTIVE SUMMARY -->
  <div style="{section_style}">
    <h2 style="{h2_style}">Executive Summary</h2>
    <p style="font-size:16px;line-height:1.7;color:#334155;margin:0">{exec_summary}</p>
    <div style="margin-top:20px">
      <div style="font-weight:700;color:#1e293b;margin-bottom:10px;font-size:14px">TRANSFORMATION THEMES</div>
      {themes_html}
    </div>
  </div>

  <!-- CURRENT STATE FINDINGS -->
  <div style="{section_style}">
    <h2 style="{h2_style}">Current State Findings</h2>
    <table style="{table_style}">
      <thead><tr>
        <th style="{th_style}">Area</th>
        <th style="{th_style}">Finding</th>
        <th style="{th_style}">Business Impact</th>
        <th style="{th_style}">Severity</th>
      </tr></thead>
      <tbody>{findings_html}</tbody>
    </table>
  </div>

  <!-- BENCHMARK GAPS -->
  <div style="{section_style}">
    <h2 style="{h2_style}">Benchmark Gaps vs. APQC Top Quartile</h2>
    <table style="{table_style}">
      <thead><tr>
        <th style="{th_style}">Metric</th>
        <th style="{th_style};text-align:center">Sprott (Est.)</th>
        <th style="{th_style};text-align:center">Top Quartile</th>
        <th style="{th_style};text-align:center">Gap</th>
      </tr></thead>
      <tbody>{benchmarks_html}</tbody>
    </table>
  </div>

  <!-- ROADMAP -->
  <div style="{section_style}">
    <h2 style="{h2_style}">Finance Transformation Roadmap</h2>
    {timeline_html}
  </div>

  <!-- INVESTMENT SUMMARY -->
  <div style="{section_style}">
    <h2 style="{h2_style}">Investment &amp; Business Case</h2>
    <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:20px">
      {inv_cards}
    </div>
    <div>
      <div style="font-weight:700;color:#1e293b;margin-bottom:8px;font-size:14px">KEY VALUE DRIVERS</div>
      <ul style="color:#475569;font-size:14px;margin:0 0 0 16px;line-height:1.8">{value_drivers}</ul>
    </div>
  </div>

  <!-- QUICK WINS -->
  <div style="{section_style}">
    <h2 style="{h2_style}">Quick Wins (90 Days)</h2>
    <table style="{table_style}">
      <thead><tr>
        <th style="{th_style}">Initiative</th>
        <th style="{th_style}">Timeline</th>
        <th style="{th_style}">Expected Savings</th>
        <th style="{th_style}">Effort</th>
      </tr></thead>
      <tbody>{qw_html}</tbody>
    </table>
  </div>

  <!-- RISKS -->
  <div style="{section_style}">
    <h2 style="{h2_style}">Key Risks &amp; Mitigations</h2>
    <table style="{table_style}">
      <thead><tr>
        <th style="{th_style}">Risk</th>
        <th style="{th_style}">Likelihood</th>
        <th style="{th_style}">Mitigation</th>
      </tr></thead>
      <tbody>{risks_html}</tbody>
    </table>
  </div>

  <!-- NEXT STEPS -->
  <div style="{section_style}">
    <h2 style="{h2_style}">Immediate Next Steps</h2>
    {next_html}
  </div>

  <!-- FOOTER -->
  <div style="text-align:center;color:#94a3b8;font-size:12px;padding:24px 0 40px">
    Confidential — prepared for Sprott Inc. by Rudra AI Finance Transformation · {today}
  </div>

</div>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run_sprott_engagement(
    input_files: list[str],
    output_path: str = "sprott_roadmap.html",
    verbose: bool = True,
) -> str:
    """
    Full pipeline: input files → HTML roadmap deliverable.

    Args:
        input_files: List of file paths (Excel, PDF, CSV).
        output_path: Where to write the HTML output.
        verbose: Print progress.

    Returns:
        Absolute path to the generated HTML file.
    """
    def log(msg: str) -> None:
        if verbose:
            print(msg)

    log(f"\n{'='*60}")
    log("  SPROTT FINANCE TRANSFORMATION ENGAGEMENT")
    log(f"{'='*60}")
    log(f"  Input files : {', '.join(Path(f).name for f in input_files)}")
    log(f"  Output      : {output_path}")
    log(f"{'='*60}\n")

    # Step 1 — extract file content
    log("Step 1/3  Extracting file content...")
    file_contents = []
    for fpath in input_files:
        log(f"  → {Path(fpath).name}")
        content = extract_file_content(fpath)
        file_contents.append(content)

    # Step 2 — Claude: analyze findings
    log("\nStep 2/3  Analyzing with Claude (diagnostic)...")
    findings = analyze_inputs(file_contents)
    log(f"  → Found {len(findings.get('process_inefficiencies', []))} process inefficiencies")
    log(f"  → Found {len(findings.get('data_quality_issues', []))} data quality issues")

    # Step 3 — Claude: generate roadmap
    log("\nStep 3/3  Generating roadmap & project plan...")
    roadmap = generate_roadmap(findings)
    phases = roadmap.get("roadmap_phases", [])
    log(f"  → {len(phases)} phases planned")
    log(f"  → Investment: {roadmap.get('investment_summary', {}).get('total_investment', '—')}")
    log(f"  → Savings: {roadmap.get('investment_summary', {}).get('annual_savings', '—')}")

    # Render HTML
    html = render_html(findings, roadmap, input_files)
    out_path = Path(output_path)
    out_path.write_text(html, encoding="utf-8")

    log(f"\n✓ Deliverable written → {out_path.resolve()}")
    return str(out_path.resolve())
